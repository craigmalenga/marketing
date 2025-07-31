"""
Report generation service
"""

import pandas as pd
import logging
from datetime import datetime
from sqlalchemy import func, and_, or_, distinct
from flask import current_app
from app import db
from models import (
    Application, FLGData, AdSpend, Product, Campaign,
    StatusMapping, FLGMetaMapping
)
import os

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Service for generating reports"""
    
    def generate_credit_performance_report(self, start_date=None, end_date=None, product_category=None):
        """Generate credit performance by product report"""
        try:
            # Build base query
            query = db.session.query(
                FLGData.product_name,
                func.count(distinct(FLGData.reference)).label('enquiry_count'),
                func.sum(FLGData.sale_value).label('enquiry_value')
            ).join(
                StatusMapping,
                FLGData.status == StatusMapping.status_name,
                isouter=True
            )
            
            # Apply date filter
            if start_date:
                query = query.filter(FLGData.received_datetime >= start_date)
            if end_date:
                query = query.filter(FLGData.received_datetime <= end_date)
            
            # Apply product category filter
            if product_category:
                products = Product.query.filter_by(category=product_category).all()
                product_names = [p.name for p in products]
                query = query.filter(FLGData.product_name.in_(product_names))
            
            # Group by product
            query = query.group_by(FLGData.product_name)
            
            results = query.all()
            
            # Now get application data
            report_data = []
            
            for product_name, enquiry_count, enquiry_value in results:
                if not product_name:
                    continue
                
                # Get applications for this product
                app_query = db.session.query(
                    func.count(distinct(Application.lead_id)).label('app_count'),
                    func.sum(Application.lead_value).label('app_value')
                ).join(
                    FLGData,
                    Application.lead_id == FLGData.reference
                ).filter(
                    FLGData.product_name == product_name
                )
                
                if start_date:
                    app_query = app_query.filter(Application.datetime >= start_date)
                if end_date:
                    app_query = app_query.filter(Application.datetime <= end_date)
                
                app_result = app_query.first()
                app_count = app_result.app_count or 0
                app_value = app_result.app_value or 0
                
                # Get processed applications
                processed_query = app_query.join(
                    StatusMapping,
                    FLGData.status == StatusMapping.status_name
                ).filter(
                    StatusMapping.is_application_processed == 1
                )
                
                processed_result = processed_query.first()
                processed_count = processed_result.app_count or 0
                processed_value = processed_result.app_value or 0
                
                # Get approved applications
                approved_query = app_query.join(
                    StatusMapping,
                    FLGData.status == StatusMapping.status_name
                ).filter(
                    StatusMapping.is_application_approved == 1
                )
                
                approved_result = approved_query.first()
                approved_count = approved_result.app_count or 0
                approved_value = approved_result.app_value or 0
                
                # Calculate metrics
                avg_credit_applied = app_value / app_count if app_count > 0 else 0
                pull_through_rate = app_count / enquiry_count if enquiry_count > 0 else 0
                processed_rate = processed_count / app_count if app_count > 0 else 0
                approval_rate = approved_count / processed_count if processed_count > 0 else 0
                avg_credit_per_enquiry = approved_value / enquiry_count if enquiry_count > 0 else 0
                
                report_data.append({
                    'product': product_name,
                    'number': enquiry_count,
                    'average_value_credit_applied': avg_credit_applied,
                    'combined_enquiry_credit_value': enquiry_value or 0,
                    'credit_for_applications': app_value,
                    'pull_through_rate': pull_through_rate,
                    'credit_for_processed_applications': processed_value,
                    'percent_applications_processed': processed_rate,
                    'credit_issued_for_approved_applications': approved_value,
                    'percent_processed_applications_issued': approval_rate,
                    'average_credit_issued_per_enquiry': avg_credit_per_enquiry
                })
            
            # Sort by product name
            report_data.sort(key=lambda x: x['product'])
            
            # Calculate totals
            totals = {
                'product': 'TOTAL',
                'number': sum(row['number'] for row in report_data),
                'combined_enquiry_credit_value': sum(row['combined_enquiry_credit_value'] for row in report_data),
                'credit_for_applications': sum(row['credit_for_applications'] for row in report_data),
                'credit_for_processed_applications': sum(row['credit_for_processed_applications'] for row in report_data),
                'credit_issued_for_approved_applications': sum(row['credit_issued_for_approved_applications'] for row in report_data)
            }
            
            # Calculate total averages
            if totals['number'] > 0:
                totals['average_value_credit_applied'] = totals['credit_for_applications'] / totals['number']
                totals['pull_through_rate'] = totals['credit_for_applications'] / totals['combined_enquiry_credit_value'] if totals['combined_enquiry_credit_value'] > 0 else 0
                totals['average_credit_issued_per_enquiry'] = totals['credit_issued_for_approved_applications'] / totals['number']
            else:
                totals['average_value_credit_applied'] = 0
                totals['pull_through_rate'] = 0
                totals['average_credit_issued_per_enquiry'] = 0
            
            if totals['credit_for_applications'] > 0:
                totals['percent_applications_processed'] = totals['credit_for_processed_applications'] / totals['credit_for_applications']
            else:
                totals['percent_applications_processed'] = 0
            
            if totals['credit_for_processed_applications'] > 0:
                totals['percent_processed_applications_issued'] = totals['credit_issued_for_approved_applications'] / totals['credit_for_processed_applications']
            else:
                totals['percent_processed_applications_issued'] = 0
            
            return {
                'rows': report_data,
                'totals': totals
            }
            
        except Exception as e:
            logger.error(f"Error generating credit performance report: {e}")
            raise
    
    def generate_marketing_campaign_report(self, start_date=None, end_date=None, campaign_name=None, ad_level=None):
        """Generate marketing campaign performance report"""
        try:
            # Get ad spend data
            spend_query = db.session.query(
                func.sum(AdSpend.spend_amount).label('total_spend')
            )
            
            if start_date:
                spend_query = spend_query.filter(AdSpend.reporting_end_date >= start_date)
            if end_date:
                spend_query = spend_query.filter(AdSpend.reporting_end_date <= end_date)
            if campaign_name:
                spend_query = spend_query.filter(AdSpend.meta_campaign_name == campaign_name)
            if ad_level:
                spend_query = spend_query.filter(AdSpend.ad_level == ad_level)
            
            total_spend = spend_query.scalar() or 0
            
            # Get enquiry data
            enquiry_query = db.session.query(
                func.count(distinct(FLGData.reference)).label('enquiry_count')
            )
            
            if start_date:
                enquiry_query = enquiry_query.filter(FLGData.received_datetime >= start_date)
            if end_date:
                enquiry_query = enquiry_query.filter(FLGData.received_datetime <= end_date)
            if campaign_name:
                enquiry_query = enquiry_query.filter(FLGData.campaign_name == campaign_name)
            
            enquiry_count = enquiry_query.scalar() or 0
            
            # Get application data with status breakdown
            status_data = []
            
            # Get all status mappings
            status_mappings = StatusMapping.query.all()
            
            for status in status_mappings:
                # Count FLG records with this status
                status_query = db.session.query(
                    func.count(distinct(FLGData.reference)).label('count'),
                    func.sum(FLGData.sale_value).label('value')
                ).filter(
                    FLGData.status == status.status_name
                )
                
                if start_date:
                    status_query = status_query.filter(FLGData.received_datetime >= start_date)
                if end_date:
                    status_query = status_query.filter(FLGData.received_datetime <= end_date)
                if campaign_name:
                    status_query = status_query.filter(FLGData.campaign_name == campaign_name)
                
                result = status_query.first()
                
                status_data.append({
                    'status': status.status_name,
                    'is_application_received': status.is_application_received,
                    'is_application_processed': status.is_application_processed,
                    'is_application_approved': status.is_application_approved,
                    'is_future': status.is_future,
                    'count': result.count or 0,
                    'value': result.value or 0
                })
            
            # Calculate summary metrics
            application_count = sum(row['count'] for row in status_data if row['is_application_received'] == 1)
            processed_count = sum(row['count'] for row in status_data if row['is_application_processed'] == 1)
            approved_count = sum(row['count'] for row in status_data if row['is_application_approved'] == 1)
            
            credit_issued = sum(row['value'] for row in status_data if row['is_application_approved'] == 1)
            
            # Calculate cost metrics
            cost_per_enquiry = total_spend / enquiry_count if enquiry_count > 0 else 0
            cost_per_application = total_spend / application_count if application_count > 0 else 0
            cost_per_approved_loan = total_spend / approved_count if approved_count > 0 else 0
            
            # Calculate approval rate
            approval_rate = approved_count / application_count if application_count > 0 else 0
            
            # Calculate average credit per approved
            avg_credit_per_approved = credit_issued / approved_count if approved_count > 0 else 0
            
            # Calculate ROI metrics
            credit_per_pound_spent = credit_issued / total_spend if total_spend > 0 else 0
            
            # Expected gross margin (43.2% based on Excel)
            expected_gross_margin = 0.432
            expected_gm_per_pound = credit_per_pound_spent * expected_gross_margin
            gm_return_per_pound = expected_gm_per_pound - 1  # Subtract the £1 spent
            
            return {
                'summary': {
                    'marketing_cost': total_spend,
                    'cost_per_enquiry': cost_per_enquiry,
                    'cost_per_application': cost_per_application,
                    'cost_per_approved_loan': cost_per_approved_loan,
                    'approval_rate': approval_rate,
                    'sum_of_credit_issued': credit_issued,
                    'average_credit_issued_per_successful_app': avg_credit_per_approved,
                    'credit_per_pound_spent': credit_per_pound_spent,
                    'expected_gross_margin_per_pound_spent': expected_gm_per_pound,
                    'gross_margin_return_per_pound_spent': gm_return_per_pound
                },
                'status_breakdown': status_data,
                'counts': {
                    'enquiries': enquiry_count,
                    'applications': application_count,
                    'processed': processed_count,
                    'approved': approved_count
                }
            }
            
        except Exception as e:
            logger.error(f"Error generating marketing campaign report: {e}")
            raise
    
    def export_credit_performance_report(self, start_date=None, end_date=None, product_category=None):
        """Export credit performance report to Excel"""
        try:
            # Generate report data
            report_data = self.generate_credit_performance_report(start_date, end_date, product_category)
            
            # Create DataFrame
            df = pd.DataFrame(report_data['rows'])
            
            # Add totals row
            totals_df = pd.DataFrame([report_data['totals']])
            df = pd.concat([df, totals_df], ignore_index=True)
            
            # Format columns
            df.columns = [
                'Product',
                'Number',
                'Average Value Credit Applied',
                'Combined Enquiry Credit Value',
                'Credit for Applications',
                'Pull Through Rate',
                'Credit for Processed Applications',
                '% Applications Processed',
                'Credit Issued for Approved Applications',
                '% Processed Applications Issued',
                'Average Credit Issued Per Enquiry'
            ]
            
            # Format percentage columns
            pct_columns = ['Pull Through Rate', '% Applications Processed', '% Processed Applications Issued']
            for col in pct_columns:
                df[col] = df[col] * 100
            
            # Create Excel file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'credit_performance_{timestamp}.xlsx'
            filepath = os.path.join(current_app.config['EXPORT_FOLDER'], filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Credit Performance', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Credit Performance']
                
                # Format cells
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Header formatting
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Total row formatting
                total_row = len(df)
                for cell in worksheet[total_row + 1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                
                # Number formatting
                for row in range(2, len(df) + 2):
                    worksheet.cell(row, 3).number_format = '#,##0.00'  # Average Value
                    worksheet.cell(row, 4).number_format = '#,##0.00'  # Combined Value
                    worksheet.cell(row, 5).number_format = '#,##0.00'  # Credit for Apps
                    worksheet.cell(row, 6).number_format = '0.00%'     # Pull Through Rate
                    worksheet.cell(row, 7).number_format = '#,##0.00'  # Processed Credit
                    worksheet.cell(row, 8).number_format = '0.00%'     # % Processed
                    worksheet.cell(row, 9).number_format = '#,##0.00'  # Approved Credit
                    worksheet.cell(row, 10).number_format = '0.00%'    # % Approved
                    worksheet.cell(row, 11).number_format = '#,##0.00' # Avg Per Enquiry
                
                # Adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting credit performance report: {e}")
            raise
    
    def export_marketing_campaign_report(self, start_date=None, end_date=None, campaign_name=None, ad_level=None):
        """Export marketing campaign report to Excel"""
        try:
            # Generate report data
            report_data = self.generate_marketing_campaign_report(start_date, end_date, campaign_name, ad_level)
            
            # Create summary DataFrame
            summary_df = pd.DataFrame([report_data['summary']])
            summary_df = summary_df.T
            summary_df.columns = ['Value']
            summary_df.index.name = 'Metric'
            
            # Create status breakdown DataFrame
            status_df = pd.DataFrame(report_data['status_breakdown'])
            
            # Create Excel file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'marketing_campaign_{timestamp}.xlsx'
            filepath = os.path.join(current_app.config['EXPORT_FOLDER'], filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Write summary sheet
                summary_df.to_excel(writer, sheet_name='Summary')
                
                # Write status breakdown sheet
                status_df.to_excel(writer, sheet_name='Status Breakdown', index=False)
                
                # Format sheets
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Format summary sheet
                summary_sheet = writer.sheets['Summary']
                for cell in summary_sheet['A']:
                    cell.font = Font(bold=True)
                
                # Format status sheet headers
                status_sheet = writer.sheets['Status Breakdown']
                for cell in status_sheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)
            
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting marketing campaign report: {e}")
            raise
    
    def generate_product_category_analysis(self, start_date=None, end_date=None, campaign_type=None):
        """Generate product category analysis by campaign"""
        try:
            # Base query joining FLG data with campaigns
            query = db.session.query(
                FLGData.campaign_name,
                FLGData.product_name,
                func.count(distinct(FLGData.reference)).label('enquiry_count')
            ).join(
                Product,
                FLGData.product_name == Product.name,
                isouter=True
            )
            
            # Apply filters
            if start_date:
                query = query.filter(FLGData.received_datetime >= start_date)
            if end_date:
                query = query.filter(FLGData.received_datetime <= end_date)
            
            # Group by campaign and product
            query = query.group_by(FLGData.campaign_name, FLGData.product_name)
            results = query.all()
            
            # Process results into campaign categories
            campaign_data = {}
            for campaign_name, product_name, count in results:
                if not campaign_name:
                    continue
                    
                # Determine campaign category
                campaign_category = self._categorize_campaign(campaign_name, campaign_type)
                if campaign_type and campaign_category != campaign_type:
                    continue
                    
                if campaign_category not in campaign_data:
                    campaign_data[campaign_category] = {
                        'total': 0,
                        'products': {},
                        'campaigns': set()
                    }
                
                campaign_data[campaign_category]['total'] += count
                campaign_data[campaign_category]['campaigns'].add(campaign_name)
                
                if product_name:
                    product_category = self._get_product_category(product_name)
                    if product_category not in campaign_data[campaign_category]['products']:
                        campaign_data[campaign_category]['products'][product_category] = 0
                    campaign_data[campaign_category]['products'][product_category] += count
            
            # Generate summary
            summary = []
            for category, data in campaign_data.items():
                row = {
                    'category': category,
                    'totalEnquiries': data['total'],
                    'tv': data['products'].get('Electronics', 0),
                    'sofas': data['products'].get('Sofa', 0),
                    'appliances': data['products'].get('Appliances', 0),
                    'electronics': data['products'].get('Electronics', 0),
                    'furniture': data['products'].get('Furniture', 0),
                    'other': data['products'].get('Other', 0)
                }
                
                # Calculate percentages
                if row['totalEnquiries'] > 0:
                    row['tvPct'] = round((row['tv'] / row['totalEnquiries']) * 100, 1)
                    row['sofasPct'] = round((row['sofas'] / row['totalEnquiries']) * 100, 1)
                    row['appliancesPct'] = round((row['appliances'] / row['totalEnquiries']) * 100, 1)
                    row['electronicsPct'] = round((row['electronics'] / row['totalEnquiries']) * 100, 1)
                    row['furniturePct'] = round((row['furniture'] / row['totalEnquiries']) * 100, 1)
                    row['otherPct'] = round((row['other'] / row['totalEnquiries']) * 100, 1)
                else:
                    row['tvPct'] = row['sofasPct'] = row['appliancesPct'] = 0
                    row['electronicsPct'] = row['furniturePct'] = row['otherPct'] = 0
                
                summary.append(row)
            
            # Generate detailed breakdown
            detailed = []
            for campaign_name in set(c for d in campaign_data.values() for c in d['campaigns']):
                # Get products for this specific campaign
                campaign_products = {}
                campaign_total = 0
                
                for cn, pn, count in results:
                    if cn == campaign_name:
                        if pn:
                            campaign_products[pn] = count
                            campaign_total += count
                
                if campaign_total > 0:
                    # Find top 2 products
                    sorted_products = sorted(campaign_products.items(), key=lambda x: x[1], reverse=True)
                    primary = sorted_products[0] if sorted_products else ('None', 0)
                    secondary = sorted_products[1] if len(sorted_products) > 1 else ('None', 0)
                    
                    detailed.append({
                        'campaignName': campaign_name,
                        'totalEnquiries': campaign_total,
                        'primaryProduct': primary[0],
                        'primaryPct': round((primary[1] / campaign_total) * 100, 1),
                        'secondaryProduct': secondary[0],
                        'secondaryPct': round((secondary[1] / campaign_total) * 100, 1),
                        'productMixRatio': f"{len(campaign_products)} products"
                    })
            
            # Generate chart data
            chartData = {
                'distribution': {
                    'labels': [row['category'] for row in summary],
                    'datasets': [
                        {
                            'label': 'TV',
                            'data': [row['tv'] for row in summary],
                            'backgroundColor': '#18124C'
                        },
                        {
                            'label': 'Sofas',
                            'data': [row['sofas'] for row in summary],
                            'backgroundColor': '#3BF7CA'
                        },
                        {
                            'label': 'Appliances',
                            'data': [row['appliances'] for row in summary],
                            'backgroundColor': '#E97132'
                        },
                        {
                            'label': 'Other',
                            'data': [row['other'] for row in summary],
                            'backgroundColor': '#A02B93'
                        }
                    ]
                },
                'effectiveness': {
                    'labels': ['TV', 'Sofas', 'Appliances', 'Electronics', 'Furniture'],
                    'datasets': [{
                        'label': 'Product Mix %',
                        'data': [20, 30, 25, 15, 10],  # Example data
                        'backgroundColor': 'rgba(59, 247, 202, 0.2)',
                        'borderColor': '#3BF7CA',
                        'pointBackgroundColor': '#3BF7CA'
                    }]
                }
            }
            
            # Generate insights
            insights = []
            
            # Find best performing campaign category
            if summary:
                best_category = max(summary, key=lambda x: x['totalEnquiries'])
                insights.append({
                    'title': 'Top Campaign Category',
                    'description': f'{best_category["category"]} campaigns generate the most enquiries',
                    'value': f'{best_category["totalEnquiries"]:,} total enquiries'
                })
            
            # Find most diverse campaign
            if detailed:
                most_diverse = max(detailed, key=lambda x: int(x['productMixRatio'].split()[0]))
                insights.append({
                    'title': 'Most Diverse Campaign',
                    'description': f'{most_diverse["campaignName"]} attracts the widest product range',
                    'value': most_diverse['productMixRatio']
                })
            
            return {
                'summary': summary,
                'detailed': sorted(detailed, key=lambda x: x['totalEnquiries'], reverse=True),
                'chartData': chartData,
                'insights': insights
            }
            
        except Exception as e:
            logger.error(f"Error generating product category analysis: {e}")
            raise

    def _categorize_campaign(self, campaign_name):
        """Categorize campaign based on name patterns"""
        if not campaign_name:
            return 'Unknown'
        
        name_lower = campaign_name.lower()
        
        if 'tv' in name_lower or 'television' in name_lower:
            return 'TV'
        elif 'sofa' in name_lower or 'furniture' in name_lower:
            return 'Sofa'
        elif 'appliance' in name_lower or 'kitchen' in name_lower:
            return 'Appliance'
        else:
            return 'General'

    def _get_product_category(self, product_name):
        """Get product category from product name"""
        product = Product.query.filter_by(name=product_name).first()
        return product.category if product else 'Other'

    
    def get_summary_statistics(self):
        """Get summary statistics for dashboard"""
        try:
            # Get total counts
            total_enquiries = FLGData.query.count()
            total_applications = Application.query.count()
            total_campaigns = Campaign.query.count()
            
            # Get this week's data
            from datetime import timedelta
            today = datetime.now().date()
            week_start = today - timedelta(days=today.weekday())
            
            week_enquiries = FLGData.query.filter(
                FLGData.received_datetime >= week_start
            ).count()
            
            week_applications = Application.query.filter(
                Application.datetime >= week_start
            ).count()
            
            # Get total spend this week
            week_spend = db.session.query(
                func.sum(AdSpend.spend_amount)
            ).filter(
                AdSpend.reporting_end_date >= week_start
            ).scalar() or 0
            
            # Get approval rate
            approved_count = db.session.query(
                func.count(distinct(FLGData.reference))
            ).join(
                StatusMapping,
                FLGData.status == StatusMapping.status_name
            ).filter(
                StatusMapping.is_application_approved == 1
            ).scalar() or 0
            
            approval_rate = approved_count / total_applications if total_applications > 0 else 0
            
            return {
                'total_enquiries': total_enquiries,
                'total_applications': total_applications,
                'total_campaigns': total_campaigns,
                'week_enquiries': week_enquiries,
                'week_applications': week_applications,
                'week_spend': week_spend,
                'approval_rate': approval_rate
            }
            
        except Exception as e:
            logger.error(f"Error getting summary statistics: {e}")
            raise